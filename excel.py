from openpyxl import Workbook
from openpyxl.drawing.image import Image
import io
import base64

class Excel:
    def __init__(self, filename):
        self.filename = filename
        
    def write_to_excel(pokemon_list, filename):
        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "ID", "Types", "Image"])

        for pokemon in pokemon_list:
            ws.append([pokemon.name, pokemon.id, ', '.join(pokemon.types), ""])
            img = Image(io.BytesIO(pokemon.image_data))
            ws.add_image(img, f"D{ws.max_row}")  # Adiciona a imagem à coluna D, na mesma linha dos outros dados            

    # Define uma altura fixa para todas as linhas após adicionar todas as imagens
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            ws.row_dimensions[row[0].row].height = 75

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 14

        wb.save(filename)
    

    def excel_to_base64(filename):
        with open(filename, "rb") as file:
            excel_data = file.read()
            excel_base64 = base64.b64encode(excel_data).decode("utf-8")
        return excel_base64